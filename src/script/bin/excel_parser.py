from src.script.bin.arguments import Arguments
import openpyxl
import pandas
import pandasql
from openpyxl.styles import Alignment, Font


def is_empty_integer(v) -> bool:
    return v is None or isinstance(v, int) or v.isdigit()


class ExcelParser(Arguments):
    def __init__(self):
        super().__init__()
        self.ACTIVITIES = ["MIC", "MBC", "MICb", "gentamicin"]
        self.ITEMS = [1, 2, 3, 1, 2, 3, 1, 2, 3, 1]
        self.COLUMNS = ["sheet", "row_id", "code", "pathogen", "activity", "item", "item_value"]
        self.COLUMNS_ERR = ["sheet", "cell", "actual value", "error_description"]
        self.ITEM_COL_OFFSET = 2
        self.MINIMAL_CODE_LEN = 6
        self.SQL = """SELECT pathogen, code, item_value, count(item_value) as cnt_item 
                FROM raw_data 
                WHERE activity = 'MIC'
                GROUP BY pathogen, code, activity, item_value
                HAVING count(item_value) > 1
            """

    def is_code(self, val) -> bool:
        v = str(val)
        res = v is not None and (v.partition("-")[1] == '-') and v.partition("-")[0].lstrip().rstrip().isdigit() and len(str(v.partition("-")[2])) > self.MINIMAL_CODE_LEN
        return res

    def get_raw_data(self, wbi: openpyxl.workbook.Workbook) -> pandas.DataFrame:
        code = ""
        activity = ""
        raw_data = pandas.DataFrame(columns=self.COLUMNS)
        self.log.info(f"There is {len(self.p.sheets)} sheet(s) selected.")
        for sheet_name in self.p.sheets:
            self.log.info(f"Building data for Excel worksheet {str(sheet_name)}.")
            for row_number in range(1, wbi[str(sheet_name)].max_row):
                lead = wbi[sheet_name].cell(row=row_number, column=2)
                if lead.value:
                    row_id = lead.value
                    if int(lead.value) == 1:
                        raw_code = str(lead.offset(row=-3, column=2).value)
                        code = (raw_code.partition("-")[2]).lstrip().rstrip()
                    pathogen = lead.offset(row=0, column=1).value
                    activity_id = 0
                    for item_id, item in enumerate(self.ITEMS):
                        if item == 1:
                            activity = self.ACTIVITIES[activity_id]
                            activity_id += 1
                        item_v = str(lead.offset(row=0, column=self.ITEM_COL_OFFSET + item_id).value)
                        raw_data.loc[len(raw_data)] = [str(sheet_name), row_id, code, pathogen, activity, item, item_v]
        return raw_data

    def approve_data(self, wbi) -> pandas.DataFrame:
        report_err = pandas.DataFrame(columns=self.COLUMNS_ERR)
        for sheet_name in self.p.sheets:
            for row_number in range(1, wbi[str(sheet_name)].max_row):
                lead = wbi[sheet_name].cell(row=row_number, column=2)
                if not is_empty_integer(lead.value):
                    report_err.loc[len(report_err)] = [str(sheet_name), f"B{row_number}", str(lead.value),
                                                       "Integer number expected"]
                elif not lead.value is None and int(lead.value) == 1:
                    try:
                        raw_code = str(lead.offset(row=-3, column=2).value)
                        if not self.is_code(raw_code):
                            report_err.loc[len(report_err)] = [str(sheet_name), f"D{row_number - 3}", str(raw_code), "The format of raw code could be '# - code'"]
                    except ValueError as e:
                        report_err.loc[len(report_err)] = [str(sheet_name), f"B{row_number}", str(lead.value), f"Wrong position of leading '1' {e}"]
        if len(report_err) > 0:
            self.log.warning("Errors found and will ge reported.")
            err = set(report_err['sheet'])
            valid = ", ".join(set(self.p.sheets) - err)
            err = ",".join(err)
            self.log.error(f"The sheets with errors: '{err}'")
            self.log.info(f"The valid sheets without errors: '{valid}'")
        return report_err

    def get_final_content(self, raw_data: pandas.DataFrame) -> pandas.DataFrame:
        limited_data: pandas.DataFrame = raw_data[['pathogen', 'code', 'activity', 'item', 'item_value']]
        build_item_value = pandasql.sqldf(self.SQL, locals())
        return build_item_value.pivot_table(values='item_value', index=['code'], columns=['pathogen'], aggfunc="first")

    def excel_final_formatting(self) -> None:
        wb = openpyxl.load_workbook(self.p.export_excel_file, read_only=False)
        ws = wb.active
        for c in ws['A1':'AA1'][0]:
            c.alignment = Alignment(textRotation=90)
            c.font = Font(bold=False)
        wb.save(filename=self.p.export_excel_file)
