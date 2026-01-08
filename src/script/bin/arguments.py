import argparse
import datetime
import os
from os import name

import openpyxl
from sqlalchemy import true
from tabulate import tabulate
import logging

def logForLevel(self, message, *args, **kwargs):
    if self.isEnabledFor(45):
        self._log(45, message, args, **kwargs)
def logToRoot(message, *args, **kwargs):
    logging.log(45, message, *args, **kwargs)

class Arguments:

    def __init__(self):
        ts = int((datetime.datetime.now()).timestamp())
        self.DATABASE = {"SQ_LITE" : "DRIVER={SQLite3 ODBC Driver};SERVER=localhost;DATABASE=test.db;Trusted_connection=yes"}
        self.EXCEL_EXTENSION = "xlsx"
        self.DEFAULT_EXPORT = os.path.join(os.getcwd(), f"export-{ts}.xlsx")
        self.TYPES_ESSAY = ["MBC", "MIC"]
        self.p = self.get_args()
        logging.addLevelName(45, 'SHOW')
        logging.basicConfig(format='%(asctime)s %(levelname)s :%(message)s',level=logging.DEBUG if self.p.verbose else 45)
        setattr(logging, "SHOW", 45)
        setattr(logging.getLoggerClass(), 'show', logForLevel)
        setattr(logging, 'show', logToRoot)
        self.log = logging.getLogger(__name__)

    def get_args(self):
        parser = argparse.ArgumentParser()
        generic  = parser.add_argument_group('Generic arguments')
        imp = parser.add_argument_group('Input data')
        raw = parser.add_argument_group('Import data to database')
        final = parser.add_argument_group('Final data file manipulation')
        generic.add_argument("-v", "--verbose", action='store_true', help="Verbose output")
        generic.add_argument("-n", "--no_question", action='store_true', help="Disable approval question")
        generic.add_argument("-D", "--dry_run", action='store_true', help="Dry run: If present -I validate Excel import data. Optional argument is file name with dry run results. If present -R provide info of raw data file (counts group counts by timestamps), use with -v")
        imp.add_argument("-i", "--import", name="imp", nargs=1, type=str, help=f"Import Excel file with data sources to database {self.DATABASE}. To check the content use parameter --dry_run. ")
        imp.add_argument("-I", "--import_show", nargs=1, type=str, help=f"Import showed data file to database {self.DATABASE}. To check the content use parameter --dry_run. ")
        imp.add_argument("-s", "--sheets", nargs='+', type=str, help="Source worksheets. If missing all worksheets will be used in  given import file or database")
        raw.add_argument("-r", "--is_range", nargs=2, help=" -d -j -g use  2 value as boundaries form ... to. The rounding i.e  2025.12 could be used")
        raw.add_argument("-l", "--is_list", nargs='+', help=" -d -j -g use values as list.")
        raw.add_argument("-S", "--show", nargs=1, help="Show database data in given list or range The rounding i.e  2025.12 could be used")
        raw.add_argument("-d", "--delete", type=str, help="Delete records with timestamps as the list or range (if -r is present)")
        raw.add_argument("-j", "--join", type=str, help="Join records with timestamps as the list or range (if -r is present). Actual timestamp as is used for joined data")
        final.add_argument("-e", "--export", nargs='*', help=f"Exported final excel file. Default value example: {self.DEFAULT_EXPORT}")
        final.add_argument("-t", "--type_essay", nargs='+', help="MIC and / or MBC, sheets in export file ")
        return parser.parse_args()

    def check_args(self):
        if self.p.import_file:
            self.p.import_file = self.p.import_file[0]
            wbi = openpyxl.load_workbook(self.p.import_file)
            sheet_ok_manes = []
            self.p.sheets = self.p.sheets if self.p.sheets else wbi.sheetnames
            for sheet_name in self.p.sheets:
                try:
                    rs = wbi[str(sheet_name)].max_row
                    sheet_ok_manes.append(sheet_name)
                    self.log.info(f"Worksheet '{sheet_name}' has {rs} rows.")
                except KeyError:
                    self.log.info(f"Worksheet '{sheet_name}' not exists. It is excluded")
                self.p.sheets = sheet_ok_manes

        self.log.info("List of variables:\n" + tabulate((dict(vars(self.p))).items(), headers=["Variable", "Value"], tablefmt="grid"))
        if (not self.p.no_question) and (not input("Do you like to proceed the task? [Y/n]") == "Y"):
            self.log.info("Script terminated by user.")
            exit(0)

    def save_file(self, save_function, file_path: str):
        try:
            save_function(file_path)
        except Exception as e:
            self.log.warning(f"Error saving file to '{file_path}': {e}")
            exit(1)

    def open_file(self, open_function, file_path: str):
        try:
            wbi = open_function(file_path)
        except FileNotFoundError as e:
            self.log.error(e)
            exit(1)
        except PermissionError as e:
            self.log.error(e)
            exit(1)
        return wbi