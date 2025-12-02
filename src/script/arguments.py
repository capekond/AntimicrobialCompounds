import argparse
import datetime
import os

import openpyxl
from tabulate import tabulate


class Arguments:

    def __init__(self):
        ts = int((datetime.datetime.now()).timestamp())
        self.p = None
        self.EXCEL_EXTENSION = "xlsx"
        self.DEFAULT_EXPORT = os.path.join(os.getcwd(), f"export-{ts}.{self.EXCEL_EXTENSION}")
        self.DEFAULT_RAW_EXPORT = os.path.join(os.getcwd(), f"export_raw-{ts}.{self.EXCEL_EXTENSION}")
        self.DEFAULT_DRY_RUN = os.path.join(os.getcwd(), f"errors-{ts}.{self.EXCEL_EXTENSION}")
        self.SUPPORTED_EXTENSIONS = ["csv", self.EXCEL_EXTENSION]

    def log(self, message: str, print_no_verbose=False) -> None:
        if print_no_verbose or self.p.verbose:
            print(message)

    def get_args(self):
        parser = argparse.ArgumentParser()
        parser.add_argument("-I", "--import_file", type=str,
                            help="Imported Excel file with data sources. The file must have expected content. To check the content use parameter --dry_run")
        parser.add_argument("-d", "--dry_run", action='store_true', help="Dry run: validate Excel import data ONLY")
        parser.add_argument("-D", "--dry_run_file", nargs='?', default=self.DEFAULT_DRY_RUN, type=str, help=f"File with dry run results.")
        parser.add_argument("-e", "--export_excel", action='store_true', help=f"Exported excel file. Default value example: {self.DEFAULT_EXPORT}")
        parser.add_argument("-E", "--export_excel_file", nargs='?', type=str, help=f"Name of exported excel file.")
        parser.add_argument("-r", "--export_raw", action='store_true', help=f"Exported data raw file. Default value example: {self.DEFAULT_RAW_EXPORT} ")
        parser.add_argument("-R", "--export_raw_file", type=str, help=f"Name of exported data raw file. Supported file extension: {', '.join(self.SUPPORTED_EXTENSIONS)}.")
        parser.add_argument("-s", "--sheets", nargs='+', type=str, help="Source worksheets. If missing all worksheets will be used.")
        parser.add_argument("-v", "--verbose", action='store_true', help="Verbose output")
        parser.add_argument("-n", "--no_question", action='store_true', help="Disable approval question")
        self.p = parser.parse_args()

    def check_args(self):
        if not self.p.import_file:
            self.log("Please add input data file -i, --import_file", True)
            exit(0)
        wbi = openpyxl.load_workbook(self.p.import_file)
        sheet_ok_manes = []
        self.p.sheets = self.p.sheets if self.p.sheets else wbi.sheetnames
        for sheet_name in self.p.sheets:
            try:
                rs = wbi[str(sheet_name)].max_row
                sheet_ok_manes.append(sheet_name)
                self.log(f"Worksheet '{sheet_name}' has {rs} rows.")
            except KeyError:
                self.log(f"Worksheet '{sheet_name}' not exists. It is excluded")
        self.p.sheets = sheet_ok_manes
        if self.p.export_raw:
            s = str(self.p.export_raw)
            if not (s[s.rfind(".") + 1:] in self.SUPPORTED_EXTENSIONS):
                self.p.export_raw = s + "." + self.EXCEL_EXTENSION
                self.p.ext = self.EXCEL_EXTENSION
                self.log(f"Wrong extension for raw export file. File name was changed to {self.p.export_raw} ")
            else:
                self.p.ext = s[s.rfind(".") + 1:]
        if not self.p.sheets:
            self.p.sheets = wbi.sheetnames
        self.log("List of variables:")
        self.log(tabulate((dict(vars(self.p))).items(), headers=["Variable", "Value"], tablefmt="grid"))
        if (not self.p.no_question) and (not input("Do you like to proceed the task? [Y/n]") == "Y"):
            self.log("Script terminated by user.")
            exit(0)
        return wbi
