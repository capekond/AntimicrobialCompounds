from src.script.arguments import Arguments
import openpyxl
import pandas
import pandasql
from openpyxl.styles import Alignment, Font


def is_empty_integer(value) -> bool:
    pass


def is_integer(value) -> bool:
    pass


def is_string(value, minimal_len: int) -> bool:
    pass


class ExcelParser(Arguments):
    def __init__(self):
        super().__init__()
        self.ACTIVITIES = ["MIC", "MBC", "MICb", "gentamicin"]
        self.ITEMS = [1, 2, 3, 1, 2, 3, 1, 2, 3, 1]
        self.COLUMNS = ["sheet", "row_id", "code", "pathogen", "activity", "item", "item_value"]
        self.ITEM_COL_OFFSET = 2
        self.MINIMAL_STRING_LEN = 6
        self.SQL = """SELECT pathogen, code, item_value, count(item_value) as cnt_item 
                FROM raw_data 
                WHERE activity = 'MIC'
                GROUP BY pathogen, code, activity, item_value
                HAVING count(item_value) > 1
            """

    def get_raw_data(self, wbi: openpyxl.workbook.Workbook) -> pandas.DataFrame:
        code = ""
        activity = ""
        raw_data = pandas.DataFrame(columns=self.COLUMNS)
        self.log(f"Total number of sheets in scope {len(self.p.sheets)}:")
        for sheet_name in self.p.sheets:
            self.log(f"Building data for Excel worksheet {str(sheet_name)}.")
            for row_number in range(1, wbi[str(sheet_name)].max_row):
                lead = wbi[sheet_name].cell(row=row_number, column=2)
                if lead.value:
                    row_id = lead.value
                    if int(lead.value) == 1:
                        raw_code = str(lead.offset(row=-3, column=2).value)
                        code = (raw_code.partition("-")[2]).lstrip().rstrip()
                    pathogen = lead.offset(row=0, column=1).value
                    activity_id = 0
                    for item_id, item in enumerate(self.ITEMS):
                        if item == 1:
                            activity = self.ACTIVITIES[activity_id]
                            activity_id += 1
                        item_value = str(lead.offset(row=0, column=self.ITEM_COL_OFFSET + item_id).value)
                        raw_data.loc[len(raw_data)] = [str(sheet_name), row_id, code, pathogen, activity, item,
                                                       item_value]
        return raw_data

    def approve_data(self) -> pandas.DataFrame:
        code = ""
        activity = ""
        report_err = pandas.DataFrame(columns=["sheet", "cell", "actual value", "error_description"])
        # if p.verbose:
        #     print(f"Total number of sheets in scope {len(p.sheets)}:")
        # for sheet_name in p.sheets:
        #     if p.verbose:
        #         print(f"\tChecking data for Excel worksheet {str(sheet_name)}.")
        #     for row_number in range(1, wbi[str(sheet_name)].max_row):
        #         lead = wbi[sheet_name].cell(row=row_number, column=2)
        #         if not is_empty_integer(lead.value):
        #
        #         if lead.value:
        #             row_id = lead.value
        #             if int(lead.value) == 1:
        #                 raw_code = str(lead.offset(row=-3, column=2).value)
        #                 if not is_string(raw_code):
        #
        #                 if
        #
        #                 code = (raw_code.partition("-")[2]).lstrip().rstrip()
        #             pathogen = lead.offset(row=0, column=1).value
        #             activity_id = 0
        #             for item_id, item in enumerate(self.ITEMS):
        #                 if item == 1:
        #                     activity = self.ACTIVITIES[activity_id]
        #                     activity_id += 1
        #                 item_value = str(lead.offset(row=0, column=self.ITEM_COL_OFFSET + item_id).value)
        return report_err

    def get_final_content(self, raw_data: pandas.DataFrame) -> pandas.DataFrame:
        limited_data: pandas.DataFrame = raw_data[['pathogen', 'code', 'activity', 'item', 'item_value']]
        build_item_value = pandasql.sqldf(self.SQL, locals())
        return build_item_value.pivot_table(values='item_value', index=['code'], columns=['pathogen'], aggfunc="first")

    def excel_final_formatting(self) -> None:
        wb = openpyxl.load_workbook(self.p.export_excel, read_only=False)
        ws = wb.active
        for c in ws['A1':'AA1'][0]:
            c.alignment = Alignment(textRotation=90)
            c.font = Font(bold=False)
        wb.save(filename=self.p.export_excel)
        self.log(f"File '{self.p.export_excel}' formatted.")
